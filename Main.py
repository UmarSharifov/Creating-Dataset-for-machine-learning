import pyodbc
import xlwt
import numpy
from scipy import stats
import re
from openpyxl import load_workbook
myServer = "DESKTOP-94D652C\SQLEXPRESS"
connection = pyodbc.connect('Driver={SQL Server};'
                            'Server='+myServer+';'
                            'Database=stud_data;'
                            'Trusted_Connection=yes;')

def getInsCode():
    cursorSelect = connection.cursor()
    cursorSelect.execute(
        'SELECT ID,INSTITUTEID '
        'FROM dbo.student WHERE DATEADMISION > 2011')
    studInst = cursorSelect.fetchall()
    current_ints = [8681,6, 6895, 8172]
    resultdict = dict()
    for i in studInst:
        for j in current_ints:
            if i[1] == j:
                resultdict[i[0]] = j
    return resultdict

def creating_dataset():
    wb = load_workbook('C:/DATASE2T.xlsx')
    sheet = wb.get_sheet_by_name('list')
    stud = list()
    for i in range(1, 32341):
        temp = list()
        temp.append(sheet.cell(row=i, column=1).value)
        temp.append(sheet.cell(row=i, column=2).value)
        temp.append(sheet.cell(row=i, column=3).value)
        temp.append(sheet.cell(row=i, column=4).value)
        temp.append(sheet.cell(row=i, column=5).value)
        temp.append(sheet.cell(row=i, column=6).value)
        temp.append(sheet.cell(row=i, column=7).value)
        temp.append(sheet.cell(row=i, column=8).value)
        temp.append(sheet.cell(row=i, column=9).value)
        temp.append(sheet.cell(row=i, column=10).value)
        temp.append(sheet.cell(row=i, column=11).value)
        temp.append(sheet.cell(row=i, column=12).value)
        temp.append(sheet.cell(row=i, column=13).value)
        temp.append(sheet.cell(row=i, column=14).value)
        temp.append(sheet.cell(row=i, column=15).value)
        temp.append(sheet.cell(row=i, column=16).value)
        temp.append(sheet.cell(row=i, column=17).value)
        temp.append(sheet.cell(row=i, column=18).value)
        temp.append(sheet.cell(row=i, column=19).value)
        stud.append(temp)
    stud_final = list()
    instcode = getInsCode()
    counter = 0
    for i in stud:
        if i[0] in instcode and i[18] != 2:
            stud_final.append(i)
    book = xlwt.Workbook(encoding="utf-8")
    sheet1 = book.add_sheet("Python Sheet 1")
    i = 0
    for row in stud_final:
        for j in range(19):
            sheet1.write(i, j, row[j])
        i = i + 1
    book.save("DATASETINST.xls")
    print("finish")

def specialization_per_intstitute():
    cursorSelect = connection.cursor()
    cursorSelect.execute(
        'SELECT SPECIALISATIONID '
        'FROM dbo.student WHERE DATEADMISION > 2011'
        'AND INSTITUTEID = 9'
        'GROUP BY SPECIALISATIONID'
    )
    data = cursorSelect.fetchall()
    for i in data:
        cursorSelect.execute('SELECT * from dbo.specialization WHERE ID = {0}'.format(i[0]))
        temp = cursorSelect.fetchall()
        print(temp[0][1])
def all_speciality():
    cursorSelect = connection.cursor()
    cursorSelect.execute(
        'SELECT * from dbo.specialization'
    )
    stud_final = cursorSelect.fetchall()
    book = xlwt.Workbook(encoding="utf-8")
    sheet1 = book.add_sheet("Python Sheet 1")
    i = 0
    # for row in stud_final:
    #     #     sheet1.write(i, 1, row[0])
    #     #     sheet1.write(i,2,row[1])
    #     #     sheet1.write(i,3,row[2])
    #     #     i = i + 1
    #     # book.save("SPECIALISATIONS.xls")
    #     # print("finish")
    for i in stud_final:
        print(i)

specialization_per_intstitute()